from datetime import date, timezone
import io
import os
import sys
from django.http import HttpResponse
from xlsxwriter.workbook import Workbook
from PIL import Image

from app.settings import BASE_DIR, TEMPLATE_TEXT


os.environ['DJANGO_SETTINGS_MODULE'] = 'test_project_settings'
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import django
django.setup()

from exel.models import Product
import openpyxl

    

def uploadFile(data):

    uploaded_file = data.get("file")
    product_name_col = int(data.get("priduct-col") ) - 1
    facturer_name_col = int(data.get("facturer-col") ) - 1
    start = int(data.get("start")) - 1
    end =  int(data.get("end")) - 1

    workbook = openpyxl.open(uploaded_file)
    sheet = workbook.active

    products_bulk_lust = list()
    ids = []
    print(start)
    print(end)

    i = - 1
    for row in sheet.iter_rows():
        i += 1
        if i >= start and i <= end:
            name_product = row[product_name_col].value
            if facturer_name_col >= 0:
                name_facturer = row[facturer_name_col].value
            else:
                name_facturer = ""

            products = Product.objects.filter(name=name_product)
            if len(products) == 0:
                print('create product ' + name_product)
                newProduct = Product()
                newProduct.name = name_product
                newProduct.facturer = name_facturer
                products_bulk_lust.append(newProduct)
            else:
                for p in products:
                    Product.objects.filter(id=p.id).update(facturer=name_facturer)
                    ids.append(p.id)
        else:
            continue
        
    Product.objects.bulk_create(products_bulk_lust)
    for p in products_bulk_lust:
        ids.append(p.id)
    return ids

        

    # for i in range(data.get("start") + 1, data.get("end") + 2):
    #     name_product = sheet.rows[product_name_col][i].value
    #     name_facturer = sheet.rows[facturer_name_col][i].value
    #     print(str(name_product) + " - ----- -" +  str(name_facturer))
    #     raise


        # products = Product.objects.filter(name=name_product)
        # if len(products) == 0:
        #     print('create product ' + name_product)
        #     newProduct = Product()
        #     newProduct.name = name_product
        #     newProduct.facturer = name_facturer
        #     products_bulk_lust.append(newProduct)
        # else:
        #     for p in products:
        #         Product.objects.filter(id=p.id).update(facturer=name_facturer)
        #         ids.append(p.id)
        # raise

    # for row in sheet.iter_rows():
    #     if row[0].data_type == 'n' and row[1].data_type == 's' and row[5].data_type == 's':
    #         products = Product.objects.filter(name=row[1].value)
    #         if len(products) == 0:
    #             print('create product ' + row[1].value)
    #             newProduct = Product()
    #             newProduct.name = row[1].value
    #             newProduct.facturer = row[5].value
    #             products_bulk_lust.append(newProduct)
    #         else:
    #             for p in products:
    #                 Product.objects.filter(id=p.id).update(facturer=row[5].value)
    #                 ids.append(p.id)
    #     else:
            # continue

    Product.objects.bulk_create(products_bulk_lust)
    for p in products_bulk_lust:
        ids.append(p.id)
    return ids

def getProducts(product_ids):
    products = []
    for id in product_ids:
        try:
            products.append(Product.objects.get(pk=id))
        except: 
            print("Продукт не найден")
    return products


def changeProdcut(data):
    product = Product.objects.get(pk=data.get('id'))
    product.name = data.get('name')
    img = data.get("file")
    print(img)
    if img != None:
        product.image = img
  
    product.place = data.get('place')
    product.facturer = data.get('facturer')
    product.facturer_сountry = data.get('facturer_сountry')
    product.descripton = data.get('descripton')
    product.save()
    return True

def createExlx(product_ids):
    output = io.BytesIO()
    workbook = Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet()
    worksheet.set_column(0, 1, 20)
    cell_format = workbook.add_format({
        'border':   1,
        'text_wrap': True
    })

    header_format = workbook.add_format({
        'border':   1,
        'text_wrap': True
    })
    cell_format.set_align('top')

    header_format.set_text_wrap()
    header_format.set_align('top')
    worksheet.set_column(2, 3, 20)
    worksheet.set_column(4, 4, 10)
    worksheet.set_column(5, 5, 20)

    merge_format = workbook.add_format({    
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
        })
    worksheet.merge_range('A1:F8', TEMPLATE_TEXT, merge_format)
    products = getProducts(product_ids)
    i = 9
    for p in products:
        j = 0
        for attr in p.getAttr():
            if j == 0 and attr != '':
                cell_width = 150
                cell_height = 110
                image_path = str(BASE_DIR) + "\\" + attr
                image = Image.open(image_path)
                h, w = image.size
                dpi = 94

                try:
                    dpi = float(image.info['dpi'][0])
                finally:
                    image.close
                cell_height = (dpi * 0.94) * (cell_height / 100)
                cell_width =  (dpi * 0.94) * (cell_width / 100)

                correction_img = 0
                if w > h:
                    correction_img = h / w
                    h = h / correction_img
                else:
                    correction_img = w / h
                    w = w / correction_img
                
                x_scale = cell_width / w
                y_scale = cell_height / h


                worksheet.write(i, j, "", cell_format)
                worksheet.insert_image(i, j, image_path, {"x_scale": x_scale, "y_scale": y_scale, "x_offset": 3, "y_offset": 3})
            else:
                worksheet.write(i, j, attr, cell_format)
            j+=1
            worksheet.set_row(i, 100) 
        i+=1
    header_format.set_align('center')
    header_format.set_align('vcenter')
    worksheet.write(8, 0, 'Фото', header_format)
    worksheet.write(8, 1, 'Наименование ', header_format)
    worksheet.write(8, 2, 'Область применения', header_format)
    worksheet.write(8, 3, 'Примечание \n (параметры)', header_format)
    worksheet.write(8, 4, 'Страна производителя', header_format)
    worksheet.write(8, 5, 'Производитель', header_format)

    workbook.close()

    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=output.xlsx"

    output.close()
    return response

